from openpyxl import load_workbook


# Load the main data file.
data_wb = load_workbook('./data.xlsx')
data_ws = data_wb.active


# Load the template file.
template_wb = load_workbook('./template.xlsx')
template_ws = template_wb.active


# Set column to template cell mapping.
column_to_template_cell = {
    # EG: The FJ column in the data file must be mapped to the F8 Cell in
    # the template file.
    'FJ': 'F8',
}


for i, row in enumerate(data_ws.rows):
    for cell in row:
        if column_to_template_cell.get(cell.column, None):
            template_ws[column_to_template_cell[cell.column]] = cell.value
    template_wb.save('row {0}.xlsx'.format(i))
